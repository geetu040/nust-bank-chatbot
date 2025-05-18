import argparse
import json
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, column_index_from_string, get_column_letter
from config import QUESTION_COLOR, ANSWER_SEP, QUESTION_SEP
from utils import clean_text

def extract_qa_from_excel(path, verbose=False):
    wb = load_workbook(path, data_only=True)
    pairs = {}

    sheetnames = wb.sheetnames
    if verbose:
        print(f"Found sheets: {sheetnames}")

    for sheet_name in tqdm(sheetnames, desc="Processing sheets"):
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows())
        for row in tqdm(rows, desc=f"Processing rows in {sheet_name}", leave=False):
            for cell in row:
                if cell.value is None:
                    continue
                if cell.fill.start_color.rgb != QUESTION_COLOR:
                    continue

                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                        n_cols_spanned = max_col - min_col + 1
                        break
                else:
                    n_cols_spanned = 1

                answers = []
                running_col = cell.column_letter
                for n_col in range(n_cols_spanned):
                    running_row = cell.row
                    while True:
                        running_row += 1
                        anwer_cell = sheet[f"{running_col}{running_row}"]
                        if anwer_cell.value is None or anwer_cell.fill.start_color.rgb == QUESTION_COLOR:
                            break
                        answers.append(clean_text(anwer_cell.value))
                    col_index = column_index_from_string(running_col)
                    running_col = get_column_letter(col_index + 1)

                if len(answers) == 0:
                    continue

                question = sheet_name + QUESTION_SEP + cell.value
                question = clean_text(question)
                answer = ANSWER_SEP.join(answers)
                pairs[question] = answer

    return pairs

"""
python prepare_excel_data.py --path "data/NUST Bank-Product-Knowledge.xlsx" --output "processed_data/excel_output.json" --verbose
"""

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--path', type=str, required=True, help="Path to Excel file")
    parser.add_argument('--output', type=str, required=True, help="Path to output JSON file")
    parser.add_argument('--verbose', action='store_true', help="Enable verbose output")
    args = parser.parse_args()

    qa_pairs = extract_qa_from_excel(args.path, args.verbose)

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(qa_pairs, f, ensure_ascii=False, indent=2)

    print(f"Extracted QA pairs saved to {args.output}")
